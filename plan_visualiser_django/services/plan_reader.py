"""
Module which reads in plan data from the file in any supported format, and
then parses the data to extract information for each activity.
"""
from abc import ABC, abstractmethod
from datetime import datetime, date
from pathlib import Path
from typing import List, Dict, Callable, Any

import openpyxl as openpyxl
from django.core.files import File

from plan_visualiser_django.models import PlanField, PlanFieldMappingType, PlanMappedField


# Utility functions for parsing into every data type required by the app from every plan format
# required to be supported.


def convert_dispatch(input_type: str, output_type: str, input_value: any) -> any:
    # Define conversion functions for each input/output type which needs to be supported
    # the hope is that this general approach will work in most cases, with some additional functions to deal with cases
    # when the simple approach doesn't work - such as if a date is encoded within a string in an unusual way.
    convert_pass_through: Callable[[Any], Any] = lambda x: x
    convert_string_int: Callable[[str], int] = lambda string: int(string.strip())
    convert_string_nnd_int: Callable[[str], int] = lambda string: int(('0'+string.strip())[:-1])
    convert_string_float: Callable[[str], float] = lambda string: float(string.strip())
    convert_string_date_dmy: Callable[[str], date] = lambda date_str: datetime.strptime(date_str, '%d:%m:%Y').date()
    convert_int_string: Callable[[int], str] = lambda int_val: str(int_val)
    convert_int_float: Callable[[int], float] = lambda int_val: float(int_val)
    convert_float_string: Callable[[float], str] = lambda float_val: str(float_val)
    convert_float_int: Callable[[float], int] = lambda float_val: int(float_val)

    convert_dispatch_table = {
        'STR': {
            'STR': convert_pass_through,
            'INT': convert_string_int,
            'DATE': convert_string_date_dmy,
        },
        'STR_nnd': {
            'INT': convert_string_nnd_int,
        },
        'INT': {
            'STR': convert_int_string,
            'INT': convert_pass_through,
        },
        'FLOAT': {
            'INT': convert_float_int,
        },
        'DATE': {
            'DATE': convert_pass_through,
        }
    }
    convert_function = convert_dispatch_table[input_type][output_type]
    converted_value = convert_function(input_value)

    return converted_value


class PlanParser():
    """
    Takes data from a plan which has already been read in into rows of data, with one row for each activity within the
    plan, and parses the columns from each activity according to the defined parsing rules, into the correct
    type and value for each field required in order to manage the plan.
    """

    def __init__(self, plan_field_mapping: PlanFieldMappingType):
        """
        plan_data will be a List of Dictionaries, with one dictionary entry for each colum in the plan file which has
        been read.

        :param plan_data:
        """
        self.column_mapping = plan_field_mapping

    def parse(self, data: List[Dict]) -> List[Dict]:
        plan_fields = PlanField.objects.all()

        parsed_data = []

        for plan_record in data:
            parsed_data_record = {}
            # Iterate through the fields for a plan map field name and input type from the plan file that has been read.
            for plan_field in plan_fields:
                field_present = True  # Adjust if there is no column for the field, and it is optional
                mapped_field_column = None
                # Should get exactly one match unless the field is optional and not included in the mapping
                try:
                    mapped_field_column = self.column_mapping.planmappedfield_set.get(
                        mapped_field__field_name=plan_field.field_name)
                except PlanMappedField.DoesNotExist as e:
                    # No mapping record for this plan field.  Ok if it's optional, otherwise an error
                    if plan_field.required_flag:
                        raise e(f"Missing compulsory mapping for field {plan_field.field_name}")
                    else:
                        field_present = False
                finally:
                    if field_present:
                        mapped_field_column_raw_value = plan_record[mapped_field_column.input_field_name]
                        field_parsed_value = convert_dispatch(mapped_field_column.input_field_type, plan_field.field_type, mapped_field_column_raw_value)
                        parsed_data_record[plan_field] = field_parsed_value
                    else:
                        parsed_data_record[plan_field] = "(n/a)"
            parsed_data.append(parsed_data_record)

        return parsed_data


class PlanFileReader(ABC):
    """
    Base abstract class which defines key operations for a given technical file format.
    This class reads in the data into a List of Dictionaries with:
    - One List entry per activity from the plan.
    - One Dictionary entry per column from the input format.
    - Where possible and appropriate, un-required columnns will be discarded.

    It then parses the data with the supplied parser which will then:
    - Extract required fields using the appropriate columns for the supplied file
    - Parse the fields into the correct type/value.

    The types of file that we will want to support (eventually) will include:
    - Excel (a different reader for each Excel version that needs to be supported).
    - CSV
    - MS Project (native (XML) - not Excel import)


    The generic flow which is defined to allow appropriate hooks
    """
    file_type_name = "(Undefined)"

    def __init__(self):
        self.parser = PlanParser

        self.pre_processing()
        self.post_processing()

    def pre_processing(self):
        """
        Allows any one off logic to be carried out before the reading of the file commences, which is
        dependent upon file type
        :return:
        """
        pass

    @abstractmethod
    def read(self, pathname: str) -> List[Dict]:
        """
        Override with code to read records from specific format
        :param pathname:
        :return:
        """
        pass

    def post_processing(self):
        pass


class ExcelXLSFileReader(PlanFileReader):
    def read(self, file_path: str) -> List[Dict]:
        """
        For now the only logic here is to hard code the sheet name that the plan is located within.  This is only
        temporary and this will have to be replaced by file type specific data fields within the database.

        :return:
        """
        self.sheet_name = Path(file_path).stem

        skiprows = 0

        wb_obj = openpyxl.load_workbook(file_path, data_only=True)
        sheet = wb_obj[self.sheet_name]
        start_row = 1 + skiprows

        table = {}
        headings = self.get_headers(sheet, start_row)
        table = {heading: [] for heading in headings}

        num_blank_rows = 0
        max_blank_rows = 100
        read_row_num = start_row
        while num_blank_rows <= max_blank_rows:
            row_confirmed = self.read_row(read_row_num, sheet, headings, table)
            if not row_confirmed:
                num_blank_rows += 1
            read_row_num += 1

        iterable_by_row = self.iterrows(table)
        return iterable_by_row

    def parse(self, raw_data, plan_field_mapping):
        parser = self.parser(plan_field_mapping=plan_field_mapping)
        return parser.parse(raw_data)

    @staticmethod
    def read_row(table_row_num, sheet, headings, table, skiprows=0):
        maybe_row = {}
        row_confirmed = False
        for col, heading in enumerate(headings):
            cell_value = sheet.cell(table_row_num + skiprows + 1, col + 1).value
            maybe_row[heading] = cell_value
            if cell_value is not None:
                row_confirmed = True
        if row_confirmed:
            for heading in headings:
                table[heading].append(maybe_row[heading])
        return row_confirmed

    @staticmethod
    def get_headers(sheet, start_row):
        headings = []  # Need to store headings in column order so list not dict
        blank = False
        row = start_row
        column = 1
        while not blank:
            maybe_heading = sheet.cell(row, column).value
            if maybe_heading is not None:
                headings.append(maybe_heading)
                column += 1
            else:
                blank = True
        return headings

    @staticmethod
    def iterrows(excel_table):
        # Work out num rows by checking any of the column arrays
        num_rows = len(next(iter(excel_table.values())))
        return [{heading: excel_table[heading][row] for heading in excel_table} for row in range(num_rows)]
